# -*- coding: utf-8 -*-
"""
Created on Sun Oct  8 16:27:15 2017

@author: Eli
"""
import openpyxl as px
import numpy as np
import datetime as dt
import pickle

#GET or create wb
class Result_Book(object):

    def __init__(self, p_dict, fold):

        self.p_dict = p_dict
        
        self.c_filename = p_dict["write_file"]
        self.filename = p_dict["aux_file_name"] + '/' + p_dict["write_file"]
        self.pic_strs = p_dict["test_file_str"]
        #print(p_dict["test_file_lab"])
        self.pic_lab = p_dict["test_file_lab"]
        temp = self.pic_lab
        self.num_pics = len(temp)
        self.fold = fold
        self.foldname = "Fold_" + str(fold)
        self.new_flag = False
        self.new_fold_flag = False
        self.wb = self.get_or_create_wb(self.filename)
        self.wfs = self.get_or_create_fold_sheet(self.foldname)
        self.wos = self.wb["Overall_Accuracy"]
        
        
        
    def checks(self):

        assert type(self.pic_strs) == list, "Image Filenames must be a list of strings"
        assert type(self.filename) == str, "Excel Filename must be a string"
        assert type(self.pic_lab) == list, "Image Labels must be a list of ints"
        assert len(self.pic_strs) == len(self.pic_lab), "Labels and Images must be same length"
        assert type(fold) == int, "Fold must be an int"

        return None

    def get_or_create_wb(self,file):
        try:
            wb = px.load_workbook(file)
#            print("Existing Workbook")
            return wb
        except FileNotFoundError:
            print("No Workbook found in directory, making one now")
            self.new_flag = True
            wb = px.Workbook()
            ws = wb.active
            ws.title = "Overall_Accuracy"
            today = dt.date(2017,11,8)
            ws['A1'] = "Date Created:"
            ws['A2'] = today.today()
            ws['A4'] = "Run Date"
            ws['B4'] = "Fold Used"
            ws['C4'] = "Accuracy"
            ws['D4'] = "Loss"
            ws['E4'] = "Total Steps"
            ws['F4'] = "Model Name"
            ws['G4'] = "Folder Used"
            ws['H4'] = "Comments"
            ws['C1'] = "RunCount:"
            ws['C2'] = 0
            wb.save(file)
            del wb, today
            wb = self.get_or_create_wb(file)
            return wb
    def get_or_create_fold_sheet(self, foldname):
        try:
            wfs = self.wb[foldname]
            return wfs
        except KeyError:
            wfs = self.wb.create_sheet(foldname)
            self.new_fold_flag = True
            wfs = self.make_new_fold_sheet(wfs)
            return wfs
    def make_new_fold_sheet(self,wfs):
        today = dt.date(2017,11,8)
        wfs['A1'] = "Date Created:"
        wfs['A2'] = today.today()
        del today
        wfs['A4'] = "Image Names"
        wfs['B4'] = "Actual Class"
        
        ind = 0
        for row in wfs.iter_rows(min_row=5, min_col=1, max_col=1, max_row=5+self.num_pics-1):
            for cell in row:
                cell.value = self.pic_strs[ind]
                ind += 1
        ind = 0
        for row in wfs.iter_rows(min_row=5, min_col=2, max_col=2, max_row=5+self.num_pics-1):
            for cell in row:
                cell.value = self.pic_lab[ind]
                ind += 1
        self.wb.save(self.filename)
        return wfs
    
    def log_run(self,pred_list=[],
                acc_dict={"accuracy": "Not Rec","loss": "Not Rec","global_step": "Not Rec"},
                model_name="Not Rec",
                folder_used="Not Rec",
                comm=''):

        p_dict = self.p_dict
        model_name = p_dict["model_file"]
        folder_used = p_dict["img_folder"]
        comm = p_dict["comment"]

        pred_list = list(pred_list)
        if pred_list:
                if len(pred_list) != self.num_pics:
                    print("Just letting you know predictions doesn't seem to have same length as the original")
        today = dt.date(2017,11,8)
        wos = self.wos
        c = wos['C2']
        old_run_count = c.value
        run_count = int(old_run_count) + 1
        wos['C2'] = run_count
        o_t_row = wos.max_row + 1
        wos.cell(row=o_t_row,column=1,value=today.today())
        wos.cell(row=o_t_row,column=2,value=self.fold)
        wos.cell(row=o_t_row,column=3,value=acc_dict["accuracy"])
        wos.cell(row=o_t_row,column=4,value=acc_dict["loss"])
        wos.cell(row=o_t_row,column=5,value=acc_dict["global_step"])
        wos.cell(row=o_t_row,column=6,value=model_name)
        wos.cell(row=o_t_row,column=7,value=folder_used)
        wos.cell(row=o_t_row,column=8,value=comm)
        self.wb.save(self.filename)
        self.wos = wos
        
        wfs = self.wfs
        
        target_col = int(wfs.max_column + 1)
        target_row = 5
        wfs.cell(row=2,column=target_col,value='RUN:')
        wfs.cell(row=3,column=target_col,value='RUNDATE:')
        wfs.cell(row=2,column=target_col+1,value=run_count)
        wfs.cell(row=3,column=target_col+1,value=today.today())
        
        
        ###Getting Column Locations###
        first_dict = pred_list[0]

        key_size = {}
        
        for key in first_dict:
            if type(first_dict[key]) == np.int64:
                key_size[key] = 1
            else:
                temp_shape= first_dict[key].shape
                key_size[key] = temp_shape[0]
#        print(key_size)
        start = target_col
        column_location = {}
        
        for key in key_size:
            end = start + key_size[key]
            column_location[key] = list(range(start,end))
            start = end
        
        ###REarrange to be a dict of lists rather than a list of dicts
        pred_dict = {}
        for key in key_size:
            temp_list = []
            for item in pred_list:
                temp_list.append(item[key])
            pred_dict[key] = temp_list
        
        ###PUt in data###
        for key in pred_dict:
            wfs.cell(row=target_row-1,column=column_location[key][0],value=key)
            #This makes act_col the column we are using
            row_move = 0
            for item in pred_dict[key]:
                if key_size[key] == 1:
#                    print("On Row",target_row+row_move,"\tOn Col",column_location[key])
                    wfs.cell(row=target_row+row_move,column=column_location[key][0],value=item)
                else:
                    for act_col in range(key_size[key]):
                        wfs.cell(row=target_row+row_move,column=column_location[key][act_col],value=item[act_col])
                row_move += 1
        self.wfs = wfs
        self.wb.save(self.filename)
        print("Logged Into",self.filename)
        self.copy_file()
        self.print2screen(pred_list)
        return None

    def print2screen(self,pred_list):

        p_dict = self.p_dict

        test_file_lab = self.pic_lab
        write_flag = p_dict["write_flag"]

        for i,p in enumerate(pred_list):
            if write_flag:
                if i % 10 == 0:
                    pred_str = 'Pred %s: %s\t%s\t%s\n' % (i + 1, p['classes'], p['probabilities'], p['logits'])
                    if p['classes'] != test_file_lab[i]:
                        pred_str = 'Wrong!!!! ' + pred_str
                    print(pred_str)

        return None

    def copy_file(self):
        wb = self.wb
        c_file = self.c_filename
        c_split = c_file.split('.')
        copy_filename = c_split[0] + "_copy.xlsx"
        wb.save(copy_filename)
        self.wb = wb    
            
#        sym_len = len(symbol)
#        total= 0
#        for i in range(sym_len):
#            total += np.power(26,i) + lis_lets.index(symbol[i])
#        return total
            
#    def write_range_on_col(self,ws,col,start_row,values):
#        if type(col) == list:
#            for c in col
#        if type(col) == str:
            
###############STARTING################
if __name__ == '__main__':
    

    xlfile = "barker.xlsx"
    fold = 4
    letters = 'abcdefghijklmnopqrstuvwxyz'
    pic_list = list(letters)
    pic_lab = [1]*len(pic_list)
    
    pred = pickle.load( open("save_pred.p","rb"))
    acc = pickle.load( open("save_eval.p","rb"))
#    acc = {"accuracy": "elevendy", "loss" : "so much", "global_step" : "42"}
    
    Result = Result_Book(xlfile,pic_list,pic_lab,fold)
    Result.log_run(pred,acc,'models','foldlets','sparklewhoof')
#    #
#    #wb = Result.wb
#    #wfs = Result.wfs
#    #
#    #
#    #print("sheetnames for wb are",wb.sheetnames)
#    #
#    #print(Result.new_flag)
#    #print(Result.new_fold_flag)
#    #print(wfs.max_column)
#    
#    
#    
#    pred_lip = {"butt": [3]*len(pic_list), "stuff": [4]*len(pic_list) }
#    
#    comment="I Likea do da cha cha"
#    
#    Result.log_run(pred_lip,acc,comment)
#    Result.copy_file()
#    
#    del Result
#    Result = Result_Book(xlfile,pic_list,pic_lab,fold+2)
#    
#    
#    pred_lip = {"butt": [3]*len(pic_list), "stuff": [4]*len(pic_list) }
#    acc = {"accuracy": "elevendy"}
#    comment="I Likea do da cha cha"
#    
#    Result.log_run(pred_lip,acc,comment)
#    Result.copy_file()
#    #wb,new_flag = get_or_create_wb(xlfile)
#    #fold_name = "Fold_" + str(fold)
#    #
#    ####Checking to see if foldsheet is made
#    
#    #    
#    #
#    #
#    #print("sheetnames for wb are",wb.sheetnames)
#    
#    
#    
#    
#    
#    #wb = Workbook()
#    #
#    #ws = wb.active
    #ws.title = "Overall_Accuracy"
